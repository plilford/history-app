"""
Import 'events and periods' tab from the source spreadsheet into Supabase.

Idempotent: re-running upserts rows, doesn't duplicate.

Usage:
    pip install -r requirements.txt
    cp ../.env.example ../.env  # then edit with your Supabase URL + service key
    python import_spreadsheet.py path/to/main_sheet.xlsx
"""

from __future__ import annotations

import os
import re
import sys
import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from supabase import create_client, Client


# ------------------------- column map ----------------------------------------
# Header row in the spreadsheet is row 3. Data starts at row 4.
HEADER_ROW = 3
DATA_START_ROW = 4

# Fixed column positions (1-indexed) from the events-and-periods tab.
COL = {
    "description":      1,    # A
    "event_short":      2,    # B
    "person":           3,    # C
    "display_name":     4,    # D
    "id":              13,    # M
    "type":            15,    # O
    "region":          16,    # P
    "country_at_time": 17,    # Q
    "country_modern":  18,    # R
    "start_year":      20,    # T
    "start_month":     21,    # U
    "start_day":       22,    # V
    "end_year":        23,    # W
    "end_month":       24,    # X
    "end_day":         25,    # Y
    "date_kind":       26,    # Z  ('period' / 'point' / etc.)
    "display_date":    28,    # AB
    "period_length":   29,    # AC
    "wikipedia_link":  30,    # AD
    "other_link":      32,    # AF
    "main_priority":   33,    # AG  ('Max priority')
    # AH (34) is computed by the DB trigger — we do NOT import it.
}

# Per-timeline priority columns start at AI = 35 and run through CQ = 95.
TIMELINE_COL_START = 35
TIMELINE_COL_END   = 95

DEFAULT_FEATURED = {
    "Worldwide: main",
    "Arts and thoughts: main",
    "UK: main",
}


# ------------------------- helpers --------------------------------------------
def slugify(s: str) -> str:
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def coerce_int(v: Any) -> int | None:
    if v is None or v == "" or v == "#N/A":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        # sometimes years are stored as strings like '1850'
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def coerce_float(v: Any) -> float | None:
    if v is None or v == "" or v == "#N/A":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def coerce_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == "#N/A":
        return None
    return s


def chunked(seq: list, n: int):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


# ------------------------- import logic ---------------------------------------
def load_timelines(ws) -> list[dict]:
    """Read the timeline names from the header row, columns AI..CQ."""
    timelines = []
    for col in range(TIMELINE_COL_START, TIMELINE_COL_END + 1):
        name = coerce_str(ws.cell(row=HEADER_ROW, column=col).value)
        if not name:
            continue
        # Skip the rare placeholder header values like 0
        timelines.append({
            "_col": col,
            "name": name,
            "slug": slugify(name),
            "display_order": col - TIMELINE_COL_START,
            "is_featured": name in DEFAULT_FEATURED,
        })
    return timelines


def load_events(ws) -> tuple[list[dict], list[dict]]:
    events: list[dict] = []
    priorities: list[dict] = []

    timeline_cols = load_timelines(ws)
    col_to_name = {t["_col"]: t["name"] for t in timeline_cols}

    seen_ids: set[int] = set()
    fallback_id = 10_000_000  # for rows missing column M

    for r in range(DATA_START_ROW, ws.max_row + 1):
        # Skip blank rows
        if not (ws.cell(row=r, column=1).value or ws.cell(row=r, column=2).value):
            continue

        eid = coerce_int(ws.cell(row=r, column=COL["id"]).value)
        if eid is None:
            fallback_id += 1
            eid = fallback_id
        # de-dup IDs (the spreadsheet has some duplicates)
        if eid in seen_ids:
            fallback_id += 1
            eid = fallback_id
        seen_ids.add(eid)

        events.append({
            "id":              eid,
            "description":     coerce_str(ws.cell(row=r, column=COL["description"]).value),
            "event_short":     coerce_str(ws.cell(row=r, column=COL["event_short"]).value),
            "person":          coerce_str(ws.cell(row=r, column=COL["person"]).value),
            "display_name":    coerce_str(ws.cell(row=r, column=COL["display_name"]).value),
            "type":            coerce_str(ws.cell(row=r, column=COL["type"]).value),
            "region":          coerce_str(ws.cell(row=r, column=COL["region"]).value),
            "country_at_time": coerce_str(ws.cell(row=r, column=COL["country_at_time"]).value),
            "country_modern":  coerce_str(ws.cell(row=r, column=COL["country_modern"]).value),
            "start_year":      coerce_int(ws.cell(row=r, column=COL["start_year"]).value),
            "start_month":     coerce_int(ws.cell(row=r, column=COL["start_month"]).value),
            "start_day":       coerce_int(ws.cell(row=r, column=COL["start_day"]).value),
            "end_year":        coerce_int(ws.cell(row=r, column=COL["end_year"]).value),
            "end_month":       coerce_int(ws.cell(row=r, column=COL["end_month"]).value),
            "end_day":         coerce_int(ws.cell(row=r, column=COL["end_day"]).value),
            "is_period":       (coerce_str(ws.cell(row=r, column=COL["date_kind"]).value) or "").lower() == "period",
            "display_date":    coerce_str(ws.cell(row=r, column=COL["display_date"]).value),
            "period_length":   coerce_int(ws.cell(row=r, column=COL["period_length"]).value),
            "wikipedia_link":  coerce_str(ws.cell(row=r, column=COL["wikipedia_link"]).value),
            "other_link":      coerce_str(ws.cell(row=r, column=COL["other_link"]).value),
            "main_priority":   coerce_float(ws.cell(row=r, column=COL["main_priority"]).value),
            # main_category is computed by the DB trigger.
        })

        # Per-timeline priorities
        for col, name in col_to_name.items():
            p = coerce_float(ws.cell(row=r, column=col).value)
            if p is None or p <= 0:
                continue
            priorities.append({
                "event_id":     eid,
                "timeline_name": name,   # resolved to ID after insertion
                "priority":     p,
            })

    return events, priorities


# ------------------------- main ----------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", help="Path to the .xlsx/.xlsm spreadsheet")
    parser.add_argument("--sheet", default="events and periods")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, don't write")
    args = parser.parse_args()

    here = Path(__file__).resolve().parent
    load_dotenv(here.parent / ".env")

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not args.dry_run:
        if not url or not key:
            sys.exit("Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env")

    print(f"Loading {args.xlsx} ...")
    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet]

    timelines = load_timelines(ws)
    print(f"  Found {len(timelines)} timelines")

    events, priorities = load_events(ws)
    print(f"  Parsed {len(events)} events, {len(priorities)} priority rows")

    if args.dry_run:
        print("Dry run; not writing.")
        return

    sb: Client = create_client(url, key)

    # 1. Upsert timelines
    print("Upserting timelines ...")
    timeline_payload = [
        {"name": t["name"], "slug": t["slug"],
         "display_order": t["display_order"], "is_featured": t["is_featured"],
         "dataset": "v1"}
        for t in timelines
    ]
    sb.table("timelines").upsert(
        timeline_payload, on_conflict="dataset,name"
    ).execute()

    # Re-read to get IDs (filter to v1 so v2 timelines don't clash)
    rows = (
        sb.table("timelines")
        .select("id,name")
        .eq("dataset", "v1")
        .execute()
        .data
    )
    name_to_id = {r["name"]: r["id"] for r in rows}

    # 2. Upsert events (dataset = 'v1') in chunks
    print("Upserting events ...")
    for ev in events:
        ev["dataset"] = "v1"
    for batch in chunked(events, 500):
        sb.table("events").upsert(batch, on_conflict="id").execute()

    # 3. Replace v1 priorities only (leave v2 priorities untouched)
    print("Replacing v1 event_timeline_priorities ...")
    v1_event_ids = [e["id"] for e in events]
    if v1_event_ids:
        # Delete in chunks; supabase has a max length on `in_` lists
        for batch in chunked(v1_event_ids, 500):
            sb.table("event_timeline_priorities").delete().in_(
                "event_id", batch
            ).execute()

    pri_payload = []
    for p in priorities:
        tid = name_to_id.get(p["timeline_name"])
        if tid is None:
            continue
        pri_payload.append({
            "event_id":    p["event_id"],
            "timeline_id": tid,
            "priority":    p["priority"],
        })

    for batch in chunked(pri_payload, 1000):
        sb.table("event_timeline_priorities").insert(batch).execute()

    print(f"Done. Inserted {len(pri_payload)} priorities. "
          f"main_category is computed by the DB trigger.")


if __name__ == "__main__":
    main()
